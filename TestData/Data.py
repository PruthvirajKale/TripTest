import openpyxl


class PageData:



    @staticmethod
    def getTestData(CityName):
        Dict = {}
        book = openpyxl.load_workbook(r"C:\\Users\\Prith\\Desktop\\Book2.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == CityName:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]

    def UrlExcel(self):
        book = openpyxl.load_workbook(r"C:\\Users\\Prith\\Desktop\\Book2.xlsx")
        sheet2 = book['Sheet2']
        URl = sheet2.cell(3, 2).value
        return self.URl
        # return sheet2.cell(3, 2).value

    print("Git Hub Practice")
    print("Git Hub Practice1")

    print("Git Hub Practice2")

    print("Git Hub Practice3")
    print("Git Hub Practice4")
    print("Git Hub Practice5")



